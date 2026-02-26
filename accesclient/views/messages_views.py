# messages_views.py
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.core.cache import cache
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import csv
from datetime import datetime, date
import json
import os
from django.conf import settings

from ..models import MessagesAscenseurs, MessagesAscenseursDetails, ArchiveMessagesAscenseurs, Appareil
from ..forms import MessageDetailForm, MessageForm


class MessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = MessagesAscenseursDetails.objects.first()
        
        # 1. Liste par défaut (comportement actuel)
        accessible_accounts = [user.first_name]

        # 2. Tentative de chargement du JSON
        json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # Si l'utilisateur est dans le fichier, on étend ses droits
                    if user.first_name in config:
                        accessible_accounts.extend(config[user.first_name])
            except Exception as e:
                print(f"Erreur lecture JSON: {e}")

        # Filter messages based on user type
        is_client = Appareil.objects.filter(Client=user.first_name).exists()
        if is_client:
            messages_list = MessagesAscenseursDetails.objects.filter(Destinataire__in=accessible_accounts)
            # For clients, derive entretiens from the messages
            entretiens = list(messages_list.values_list('entretien', flat=True).distinct())
        else:
            # For maintenance users, show messages where entretien matches OR is null/empty
            messages_list = MessagesAscenseursDetails.objects.filter(
                Q(entretien__in=accessible_accounts) | 
                Q(entretien__isnull=True) | 
                Q(entretien='')
            )
            # For maintenance users, use the accessible_accounts list directly
            # This ensures the dropdown appears even if there are no active messages for some agencies
            entretiens = sorted(list(set(accessible_accounts)))
        
        # Get the selected "Entretien" from GET parameters
        selected_entretien = request.GET.get('entretien')

        # Filter messages based on selected "Entretien"
        if selected_entretien:
            messages_list = messages_list.filter(entretien=selected_entretien)
        
        # Order by date descending (most recent first)
        messages_list = messages_list.order_by('-Date')
        
        # Pagination
        paginator = Paginator(messages_list, 50)  # Show 50 messages per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        excluded_columns = ['Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages']  
        custom_column_names = {
            'entretien': 'Agence',
            'Date': 'Date du message',
            'Nature_de_l_appel': 'Type d\'appel',
            'code_client': 'N°APP',
            'Adresse': 'Adresse',
            'Code_Postal': 'Code Postal',
            'ville': 'Ville',
            'Résidence': 'Résidence',
            'Message': 'Message',
            'Action': 'Action',
            'Nom': 'Nom',
            'Société_de_l_appelant': 'Coord. de l\'appelant',
            'Nom_de_l_appelant': 'Nom de l\'appelant',
            'Téléphone_de_l_appelant': 'Téléphone ',
            'Adresse_de_l_appelant': 'Adresse de l\'appelant',
            'Code_postal_de_l_appelant': 'Code postal de l\'appelant',
            'Ville_de_l_appelant': 'Ville de l\'appelant',
        }

        # Get Résidence from Appareil model
        for message in page_obj:
            try:
                appareil = Appareil.objects.get(N_ID=message.N_ID)
                message.Résidence = appareil.Résidence if appareil.Résidence else "--"
            except Appareil.DoesNotExist:
                message.Résidence = "--"
        
        selected_columns = [field_name for field_name in messages.get_fields() if request.GET.get(field_name)]

        return render(request, 'accesclient/mesasc.html', {
            'messages': messages,
            'messages_list': page_obj,
            'page_obj': page_obj,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'entretiens': entretiens,
            'selected_entretien': selected_entretien,
        })


class ArchiveMessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = ArchiveMessagesAscenseurs.objects.first()

        # Retrieve filter parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        selected_entretien = request.GET.get('entretien')
        search_query = request.GET.get('search', '')
        
        # Debug: print received parameters
        print(f"=== Archive Messages Filters ===")
        print(f"User: {user.first_name}")
        print(f"Start date: {start_date_str}")
        print(f"End date: {end_date_str}")
        print(f"Entretien: {selected_entretien}")
        print(f"Search: {search_query}")
        
        # 1. Get accessible accounts
        accessible_accounts = [user.first_name]
        json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    if user.first_name in config:
                        accessible_accounts.extend(config[user.first_name])
            except Exception as e:
                print(f"Erreur lecture JSON: {e}")

        # 2. Get entretiens list (always needed for dropdown)
        is_client = Appareil.objects.filter(Client=user.first_name).exists()
        if is_client:
            # For clients, get entretiens from all available messages
            all_messages = ArchiveMessagesAscenseurs.objects.filter(Destinataire__in=accessible_accounts)
            entretiens = list(all_messages.values_list('entretien', flat=True).distinct())
        else:
            # For maintenance users, use accessible_accounts to ensure dropdown always appears
            entretiens = sorted(list(set(accessible_accounts)))
        
        # 3. Fetch messages for the user based on user type
        if is_client:
            messages_list = ArchiveMessagesAscenseurs.objects.filter(Destinataire__in=accessible_accounts)
        else:
            messages_list = ArchiveMessagesAscenseurs.objects.filter(entretien__in=accessible_accounts)
        
        # 4. Apply date range filter
        if start_date_str and end_date_str:
            start_date = parse_date(start_date_str)
            end_date = parse_date(end_date_str)
            print(f"Date filter: {start_date_str} to {end_date_str}")
        elif start_date_str:
            start_date = parse_date(start_date_str)
            end_date = timezone.now()
            print(f"Date filter: {start_date_str} to now")
        elif end_date_str:
            start_date = timezone.make_aware(datetime.datetime.min)
            end_date = parse_date(end_date_str)
            print(f"Date filter: beginning to {end_date_str}")
        else:
            # Default to the last 24 hours
            start_date = timezone.now() - timezone.timedelta(days=1)
            end_date = timezone.now()
            print(f"Date filter: default 24h")

        messages_list = messages_list.filter(Date__range=[start_date, end_date])
        print(f"Messages after date filter: {messages_list.count()}")

        # 5. Filter messages based on selected "Entretien"
        if selected_entretien:
            messages_list = messages_list.filter(entretien=selected_entretien)
            print(f"Filtering by entretien: {selected_entretien}, count: {messages_list.count()}")

        # 6. Search functionality
        if search_query:
            search_filter = Q()
            for field in ArchiveMessagesAscenseurs._meta.fields:
                search_filter |= Q(**{f"{field.name}__icontains": search_query})
            messages_list = messages_list.filter(search_filter)
            print(f"Messages after search '{search_query}': {messages_list.count()}")

        # 7. Order by date descending (most recent first)
        messages_list = messages_list.order_by('-Date')

        # 8. Pagination
        paginator = Paginator(messages_list, 150)  # Show 150 messages per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Print debugging information
        print(f"Total messages after filtering: {paginator.count}")
        print(f"Messages on current page: {page_obj.paginator.per_page}")

        excluded_columns = [
            'Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 
            'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 
            'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages',
            'Adresse', 'Code_Postal', 'ville',  # Hide these as they're combined in Résidence
        ]

        # Generate custom column names dynamically in specific order
        custom_column_names = {}
        # Add columns in desired order
        for field in ArchiveMessagesAscenseurs._meta.get_fields():
            if field.name not in excluded_columns:
                custom_column_names[field.name] = field.verbose_name
                # Insert Résidence right after Date
                if field.name == 'Date':
                    custom_column_names['Résidence'] = 'Coordonnées du Site'
        
        custom_column_names['entretien'] = 'Agence'
        selected_columns = [field.name for field in ArchiveMessagesAscenseurs._meta.get_fields() if request.GET.get(field.name)]
        
        # Combine the content to create 'Résidence' for archive messages
        for message in page_obj:
            try:
                appareil = Appareil.objects.get(N_ID=message.N_ID)
                message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}, {appareil.Résidence}"
            except Appareil.DoesNotExist:
                message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}"

        # Handle export functionality
        if 'export' in request.GET:
            # For export, use all non-excluded columns
            export_columns = [field.name for field in ArchiveMessagesAscenseurs._meta.get_fields() 
                            if field.name not in excluded_columns]
            # Add 'Résidence' and make sure 'entretien' is included
            if 'entretien' not in export_columns:
                export_columns.append('entretien')
            export_columns.append('Résidence')
            
            # Prepare messages with Résidence field
            export_list = messages_list
            for message in export_list:
                try:
                    appareil = Appareil.objects.get(N_ID=message.N_ID)
                    message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}, {appareil.Résidence}"
                except Appareil.DoesNotExist:
                    message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}"
            
            return self.export_to_csv(export_list, export_columns, custom_column_names)

        return render(request, 'accesclient/archive_messages.html', {
            'messages': messages,
            'page_obj': page_obj,
            'messages_list': page_obj,  # Pass the paginated page object
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'start_date': start_date_str or '',
            'end_date': end_date_str or '',
            'entretiens': entretiens,
            'selected_entretien': selected_entretien or '',
            'search_query': search_query or '',
        })

    def export_to_csv(self, messages_list, export_columns, custom_column_names):
        """Export messages to Excel format instead of CSV"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Archive Messages"
        
        # Write header with custom names
        headers = [custom_column_names.get(col, col) for col in export_columns]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_num, message in enumerate(messages_list, start=2):
            for col_num, col in enumerate(export_columns, start=1):
                value = getattr(message, col, None)
                
                # Handle datetime fields
                if isinstance(value, datetime):
                    if timezone.is_aware(value):
                        value = timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    value = value.strftime('%Y-%m-%d')
                elif value is None:
                    value = ''
                else:
                    value = str(value)
                
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="archive_messages.xlsx"'
        
        return response


def messages_list(request):
    user = request.user
    messages = MessagesAscenseursDetails.objects.filter(Destinataire=user.username)
    
    date_str = request.GET.get('date')
    if date_str:
        try:
            date_obj = parse_date(date_str)
            if date_obj:
                messages = messages.filter(Date__date=date_obj)
        except ValueError:
            pass
    
    return render(request, 'accesclient/messages_list.html', {'messages': messages})


def message_detail(request, pk):
    message = get_object_or_404(MessagesAscenseursDetails, pk=pk)
    
    if request.method == 'POST':
        form = MessageDetailForm(request.POST)
        if form.is_valid():
            fields = [field for field, value in form.cleaned_data.items() if value]
    else:
        form = MessageDetailForm()
        fields = [field.verbose_name for field in MessagesAscenseursDetails._meta.get_fields() if field.name != 'N_des_messages']

    return render(request, 'accesclient/message_detail.html', {'message': message, 'form': form, 'fields': fields})


def create_message(request, N_ID):
    appareil = get_object_or_404(Appareil, N_ID=N_ID)
    code_client = appareil.Code_Client
    destinataire = appareil.Destinataire

    if request.method == 'POST':
        form = MessageForm(request.POST, user=request.user)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            MessagesAscenseurs.objects.create(
                Nature_de_l_appel=cleaned_data['Nature_de_l_appel'],
                Nom_de_l_appelant=cleaned_data['Nom_de_l_appelant'],
                Société_de_l_appelant=cleaned_data['Société_de_l_appelant'],
                Téléphone_de_l_appelant=cleaned_data['Téléphone_de_l_appelant'],
                Message=cleaned_data['Message'],
                Destinataire=destinataire,
                N_ID=N_ID  # Use the N_ID from URL parameter directly
            )
            return redirect('appareil_list')  
    else:
        # Initialize form with N_ID value
        form = MessageForm(user=request.user, initial={'N_ID': N_ID})

    return render(request, 'accesclient/create_message.html', {
        'form': form,
        'code_client': code_client,
        'destinataire': destinataire,
    })


def export_messages_to_excel(request):
    user = request.user
    
    # 1. Liste par défaut (comportement actuel)
    accessible_accounts = [user.first_name]

    # 2. Tentative de chargement du JSON
    json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if user.first_name in config:
                    accessible_accounts.extend(config[user.first_name])
        except Exception:
            pass

    # Filter messages based on user type using MessagesAscenseursDetails to get 'entretien'
    is_client = Appareil.objects.filter(Client=user.first_name).exists()
    if is_client:
        messages = MessagesAscenseursDetails.objects.filter(Destinataire__in=accessible_accounts)
    else:
        messages = MessagesAscenseursDetails.objects.filter(entretien__in=accessible_accounts)

    # Apply Entretien filter if present
    selected_entretien = request.GET.get('entretien')
    if selected_entretien:
        messages = messages.filter(entretien=selected_entretien)

    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    # Define headers for the Excel file
    headers = [
        'Destinataire',
        'Date',
        'Message',
        'Nom',
        'Téléphone',
        'Digicode',
        'Action',
        'Société de l\'appelant',
        'Nom de l\'appelant',
        'Adresse de l\'appelant',
        'Code postal de l\'appelant',
        'Ville de l\'appelant',
        'Téléphone de l\'appelant',
        'Digicode de l\'appelant',
        'Nature de l\'appel',
        'Agence' # Added Agence column
    ]

    # Write headers to the first row in the worksheet
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, message in enumerate(messages, 2):  # Start from row 2 for data
        ws.cell(row=row_num, column=1, value=message.Destinataire)
        # Handle datetime - check if it's timezone-aware or naive
        if message.Date:
            if timezone.is_aware(message.Date):
                local_date = timezone.localtime(message.Date)
                ws.cell(row=row_num, column=2, value=local_date.replace(tzinfo=None))
            else:
                # Naive datetime - use as-is
                ws.cell(row=row_num, column=2, value=message.Date)
        else:
            ws.cell(row=row_num, column=2, value="")
            
        ws.cell(row=row_num, column=3, value=message.Message)
        ws.cell(row=row_num, column=4, value=message.Nom)
        ws.cell(row=row_num, column=5, value=message.Téléphone)
        ws.cell(row=row_num, column=6, value=message.Digicode)
        ws.cell(row=row_num, column=7, value=message.Action)
        ws.cell(row=row_num, column=8, value=message.Société_de_l_appelant)
        ws.cell(row=row_num, column=9, value=message.Nom_de_l_appelant)
        ws.cell(row=row_num, column=10, value=message.Adresse_de_l_appelant)
        ws.cell(row=row_num, column=11, value=message.Code_postal_de_l_appelant)
        ws.cell(row=row_num, column=12, value=message.Ville_de_l_appelant)
        ws.cell(row=row_num, column=13, value=message.Téléphone_de_l_appelant)
        ws.cell(row=row_num, column=14, value=message.Digicode_de_l_appelant)
        ws.cell(row=row_num, column=15, value=message.Nature_de_l_appel)
        ws.cell(row=row_num, column=16, value=message.entretien) # Added Agence value

    filename = f"messages_{user.username}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
