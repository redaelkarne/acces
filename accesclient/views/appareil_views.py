# appareil_views.py
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q
from openpyxl import Workbook
from io import BytesIO
import json
import os
from django.conf import settings

from ..models import Appareil
from ..forms import AppareilModificationForm


class AppareilView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user

        # 1. Liste par défaut (comportement actuel)
        accessible_accounts = [user.username, user.first_name]

        # 2. Get delegated users from Appareil
        if Appareil.objects.filter(Client=user.first_name).exists():
            delegated_users = list(Appareil.objects.filter(
                Client=user.first_name
            ).values_list('Entretien', flat=True).distinct())
            # Filter out None and PERDU values
            delegated_users = [entretien for entretien in delegated_users if entretien and entretien != 'PERDU']
            accessible_accounts.extend(delegated_users)

        # 3. Tentative de chargement du JSON
        json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # Si l'utilisateur est dans le fichier, on étend ses droits
                    if user.first_name in config:
                        accessible_accounts.extend(config[user.first_name])
            except Exception as e:
                print(f"Erreur lecture JSON: {e}")

        # Remove duplicates and PERDU
        accessible_accounts = [acc for acc in set(accessible_accounts) if acc != 'PERDU']

        # Fetch Appareil records based on user type
        # Check if the user exists as Client in Appareil table
        if Appareil.objects.filter(Client=user.first_name).exists():
            appareils_list = Appareil.objects.filter(Destinataire__in=accessible_accounts)
        else:
            appareils_list = Appareil.objects.filter(Entretien__in=accessible_accounts)

        # Use accessible_accounts directly as entretiens for filter
        entretiens = sorted(accessible_accounts)
        
        # Get the selected "Entretien" from GET parameters
        selected_entretien = request.GET.get('entretien')

        # Filter messages based on selected "Entretien"
        if selected_entretien:
            appareils_list = appareils_list.filter(Entretien=selected_entretien)

        # Example of excluding specific columns
        excluded_columns = ['s_Lineage', 'dateImport', 'Autres_1', 'Autres_2', 'Observations', 'Id_Societe']

        # Example of custom column names
        custom_column_names = {
            'Code_Client': 'Client Code',
            'DateCreation': 'Date of Creation',
            'Client': 'Client',
            'Entretien': 'Agence',
            'Destinataire': 'Destinataire',
            'Adresse': 'Adresse',
            'Code_Postal': 'Code_Postal',
            'Ville': 'Ville',
            'Résidence': 'Residence',
            'Informations': 'Informations',
            'Type': 'Type',
            'Incarcération': 'Incarcération',
            'Consigne_volatile': 'Consigne volatile',
            'Phonie': 'Phonie',
            'Transmetteur': 'Transmetteur'
        }

        # Example of dynamically fetching selected columns from request.GET
        selected_columns = [field.name for field in Appareil._meta.fields if request.GET.get(field.name)]
        search_query = request.GET.get('search', '')

        # If there's a search query, filter the appareils_list
        if search_query:
            search_filter = Q()
            for field in Appareil._meta.fields:
                search_filter |= Q(**{f"{field.name}__icontains": search_query})
            appareils_list = appareils_list.filter(search_filter)

        # Paginate the appareils_list
        paginator = Paginator(appareils_list, 20)  # Show 20 appareils per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, 'accesclient/basededonnees.html', {
            'page_obj': page_obj,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'entretiens': entretiens,
            'selected_entretien': selected_entretien,
            'search_query': search_query
        })


def modify_appareil(request, id):
    appareil = get_object_or_404(Appareil, pk=id)
    user = request.user
    
    # Get accessible accounts
    accessible_accounts = [user.username, user.first_name]
    
    if Appareil.objects.filter(Client=user.first_name).exists():
        delegated_users = list(Appareil.objects.filter(
            Client=user.first_name
        ).values_list('Entretien', flat=True).distinct())
        # Filter out None and PERDU values
        delegated_users = [entretien for entretien in delegated_users if entretien and entretien != 'PERDU']
        accessible_accounts.extend(delegated_users)
    
    json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if user.first_name in config:
                    accessible_accounts.extend(config[user.first_name])
        except Exception as e:
            print(f"Erreur lecture JSON: {e}")
    
    # Remove duplicates and PERDU
    accessible_accounts = [acc for acc in set(accessible_accounts) if acc != 'PERDU']
    # Remove PERDU from accessible accounts
    accessible_accounts = [acc for acc in accessible_accounts if acc != 'PERDU']
    
    # Get distinct clients for the dropdown
    if Appareil.objects.filter(Client=user.first_name).exists():
        clients = Appareil.objects.filter(
            Destinataire__in=accessible_accounts
        ).values_list('Client', flat=True).distinct()
    else:
        clients = Appareil.objects.filter(
            Entretien__in=accessible_accounts
        ).values_list('Client', flat=True).distinct()
    
    clients = [c for c in clients if c and c != 'PERDU']
    
    # Get distinct types
    if Appareil.objects.filter(Client=user.first_name).exists():
        types = Appareil.objects.filter(
            Destinataire__in=accessible_accounts
        ).values_list('Type', flat=True).distinct()
    else:
        types = Appareil.objects.filter(
            Entretien__in=accessible_accounts
        ).values_list('Type', flat=True).distinct()
    
    types = [t for t in types if t and t != 'PERDU']
    
    if request.method == 'POST':
        form = AppareilModificationForm(
            request.POST, 
            instance=appareil,
            clients=clients,
            entretiens=accessible_accounts,
            types=types
        )
        if form.is_valid():
            form.save()
            return redirect('appareil_list')  
    else:
        form = AppareilModificationForm(
            instance=appareil,
            clients=clients,
            entretiens=accessible_accounts,
            types=types
        )
    
    return render(request, 'accesclient/modify_appareil.html', {'form': form})


def create_appareil(request):
    """Create a new appareil with empty form fields"""
    user = request.user
    
    # Get accessible accounts (same logic as AppareilView)
    accessible_accounts = [user.username, user.first_name]
    
    # Get the client for the user - check if user is a Client
    user_client = None
    if Appareil.objects.filter(Client=user.first_name).exists():
        # User is a client, use user.first_name as the client
        user_client = user.first_name
        delegated_users = list(Appareil.objects.filter(
            Client=user.first_name
        ).values_list('Entretien', flat=True).distinct())
        # Filter out None and PERDU values
        delegated_users = [entretien for entretien in delegated_users if entretien and entretien != 'PERDU']
        accessible_accounts.extend(delegated_users)
    else:
        # User is an Entretien, find their Client
        client_record = Appareil.objects.filter(Entretien=user.first_name).first()
        if client_record:
            user_client = client_record.Client
    
    json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if user.first_name in config:
                    accessible_accounts.extend(config[user.first_name])
        except Exception as e:
            print(f"Erreur lecture JSON: {e}")
    
    accessible_accounts = list(set(accessible_accounts))
    # Remove PERDU from accessible accounts
    accessible_accounts = [acc for acc in accessible_accounts if acc != 'PERDU']
    
    # Get distinct clients for the dropdown
    if Appareil.objects.filter(Client=user.first_name).exists():
        clients = Appareil.objects.filter(
            Destinataire__in=accessible_accounts
        ).values_list('Client', flat=True).distinct()
    else:
        clients = Appareil.objects.filter(
            Entretien__in=accessible_accounts
        ).values_list('Client', flat=True).distinct()
    
    clients = [c for c in clients if c and c != 'PERDU']
    
    # Get distinct types for the user's appareils
    if Appareil.objects.filter(Client=user.first_name).exists():
        types = Appareil.objects.filter(
            Destinataire__in=accessible_accounts
        ).values_list('Type', flat=True).distinct()
    else:
        types = Appareil.objects.filter(
            Entretien__in=accessible_accounts
        ).values_list('Type', flat=True).distinct()
    
    types = [t for t in types if t and t != 'PERDU']
    
    if request.method == 'POST':
        form = AppareilModificationForm(
            request.POST, 
            clients=clients,
            entretiens=accessible_accounts,
            types=types
        )
        if form.is_valid():
            new_appareil = form.save(commit=False)
            new_appareil.Opérateur = request.user.username
            new_appareil.save()
            return redirect('appareil_list')  
    else:
        # Create form with initial client value
        form = AppareilModificationForm(
            clients=clients,
            entretiens=accessible_accounts,
            types=types,
            initial={'Client': user_client} if user_client else {}
        )
    
    return render(request, 'accesclient/modify_appareil.html', {'form': form, 'is_creating': True})


def set_appareil_perdu(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Set the specified fields to "PERDU"
    appareil.Client = "PERDU"
    appareil.Opérateur = "PERDU"
    appareil.Entretien = "PERDU"
    appareil.Destinataire = "PERDU"
    appareil.Informations = "PERDU"

    # Save the changes to the database
    appareil.save()

    # Redirect to the appareils list (or wherever you want)
    return redirect('appareil_list')


def modify_autres_if_meditrax(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Redirect if the Client is not Meditrax
    if appareil.Client != "MEDITRAX":
        return redirect('appareil_list')

    if request.method == 'POST':
        form = AppareilModificationForm(request.POST, instance=appareil)
        if form.is_valid():
            # Save only the 'Autres' field
            appareil.Autres = form.cleaned_data['Autres_1']
            appareil.save()
            return redirect('appareil_list')
    else:
        form = AppareilModificationForm(instance=appareil)

    return render(request, 'accesclient/Meditrax.html', {'form': form, 'appareil': appareil})


def sanitize_text(text):
    if text:
        return ''.join(c for c in text if c.isprintable())
    return text


def generate_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Appareils"
    user = request.user

    headers = [
        'N° ID',
        'DateCreation',
        'Client',
        'Opérateur',
        'Entretien',
        'Destinataire',
        'Adresse',
        'Code Postal',
        'Ville',
        'Résidence',
        'Informations',
        'Code Client',
        'Type',
        'Incarcération',
        'Code consigne',
        'Consigne volatile',
        'Utilisateur',
        'dateImport',
        'MES',
        'RES',
        'Phonie',
        'Transmetteur',
        'Observations',
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Logic copied from AppareilView.get to ensure consistency
    accessible_accounts = [user.first_name]
    
    # Get delegated users from Appareil
    if Appareil.objects.filter(Client=user.first_name).exists():
        delegated_users = list(Appareil.objects.filter(
            Client=user.first_name
        ).values_list('Entretien', flat=True).distinct())
        # Filter out None and PERDU values
        delegated_users = [entretien for entretien in delegated_users if entretien and entretien != 'PERDU']
        accessible_accounts.extend(delegated_users)
    
    json_path = os.path.join(settings.BASE_DIR, 'access_config.json')
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if user.first_name in config:
                    accessible_accounts.extend(config[user.first_name])
        except Exception:
            pass

    # Remove duplicates and PERDU
    accessible_accounts = [acc for acc in set(accessible_accounts) if acc != 'PERDU']

    if Appareil.objects.filter(Client=user.first_name).exists():
        appareils = Appareil.objects.filter(Destinataire__in=accessible_accounts)
    else:
        appareils = Appareil.objects.filter(Entretien__in=accessible_accounts)

    # Apply filters
    selected_entretien = request.GET.get('entretien')
    if selected_entretien:
        appareils = appareils.filter(Entretien=selected_entretien)
        
    search_query = request.GET.get('search', '')
    if search_query:
        search_filter = Q()
        for field in Appareil._meta.fields:
            search_filter |= Q(**{f"{field.name}__icontains": search_query})
        appareils = appareils.filter(search_filter)

    for row_num, appareil in enumerate(appareils.iterator(), 2):
        # Assigner chaque colonne en respectant l'ordre des headers
        col = 1
        ws.cell(row=row_num, column=col, value=appareil.N_ID); col += 1
        
        if appareil.DateCreation:
            local_date_creation = timezone.localtime(appareil.DateCreation).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col, value=local_date_creation); col += 1
        else:
            col += 1
            
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Client)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Opérateur)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Entretien)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Destinataire)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Adresse)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Code_Postal)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Ville)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Résidence)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Informations)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Code_Client)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Type)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Incarcération)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Code_consigne)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Consigne_volatile)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Utilisateur)); col += 1
        
        if appareil.dateImport:
            local_date_import = timezone.localtime(appareil.dateImport).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col, value=local_date_import); col += 1
        else:
            col += 1
            
        if appareil.MES:
            local_mes = timezone.localtime(appareil.MES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col, value=local_mes); col += 1
        else:
            col += 1
            
        if appareil.RES:
            local_res = timezone.localtime(appareil.RES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col, value=local_res); col += 1
        else:
            col += 1
            
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Phonie)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Transmetteur)); col += 1
        ws.cell(row=row_num, column=col, value=sanitize_text(appareil.Observations))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_appareils_to_excel(request):
    user = request.user
    filename = f"appareils_{user.username}.xlsx"

    response = StreamingHttpResponse(
        generate_excel(request),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
