from django.views import View
from .models import MessagesAscenseurs,MessagesAscenseursDetails , Astreinte
from django.shortcuts import render,redirect, get_object_or_404
from django.utils.dateparse import parse_date
from .forms import MessageDetailForm
from django.urls import reverse_lazy , reverse
from django.views import generic
from .forms import CustomUserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.dateparse import parse_date
from django.utils import timezone
from .models import Appareil  
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.dateparse import parse_date
from .models import ArchiveMessagesAscenseurs
import csv
from datetime import datetime
from django.core.cache import cache
from .forms import MessageForm,AppareilModificationForm,AstreinteForm
from django.http import StreamingHttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import redirect
from .forms import ExcelUploadForm
from .forms import process_excel_file
from django.contrib import messages

class MessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = MessagesAscenseursDetails.objects.first()
        messages_list = MessagesAscenseursDetails.objects.filter(Destinataire=user.username)
        
        # Get unique "Entretien" values for the user
        entretiens = messages_list.values_list('entretien', flat=True).distinct()
        
        # Get the selected "Entretien" from GET parameters
        selected_entretien = request.GET.get('entretien')

        # Filter messages based on selected "Entretien"
        if selected_entretien:
            messages_list = messages_list.filter(entretien=selected_entretien)
        
        excluded_columns = ['Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages']  
        custom_column_names = {
            'Date': 'Date du message',
            'Nature_de_l_appel': 'Type d\'appel',
            #'Destinataire': 'Destinataire',
            'code_client': 'N°APP',
            'Résidence': 'Coordonnées du Site',  # New combined column
            #'entretien': 'Entretien',
            'Message': 'Message',
            'Action': 'Action',
            'Nom': 'Nom',
            #'Téléphone': 'Téléphone',
            #'Digicode': 'Digicode',
            
            'Société_de_l_appelant': 'Coord. de l\'appelant',
            'Nom_de_l_appelant': 'Nom de l\'appelant',
            'Téléphone_de_l_appelant': 'Téléphone ',
            'Adresse_de_l_appelant': 'Adresse de l\'appelant',
            'Code_postal_de_l_appelant': 'Code postal de l\'appelant',
            'Ville_de_l_appelant': 'Ville de l\'appelant',
            
            #'Digicode_de_l_appelant': 'Digicode de l\'appelant',
            
        }

        # Combine the content of the three columns into one called 'Résidence'
        for message in messages_list:
            message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}"
        
        selected_columns = [field_name for field_name in messages.get_fields() if request.GET.get(field_name)]

        return render(request, 'accesclient/mesasc.html', {
            'messages': messages,
            'messages_list': messages_list,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'entretiens': entretiens,
            'selected_entretien': selected_entretien,
        })

def messages_list(request):
    user = request.user
    messages = MessagesAscenseursDetails.objects.filter(Destinataire=user.username)
    
    date_str = request.GET.get('date')
    if date_str:
        try:
            date_obj = parse_date(date_str)
            if date_obj:
                messages = messages.filter(Date__date=date_obj)
        except ValueError:
            pass
    
    return render(request, 'accesclient/messages_list.html', {'messages': messages})

def message_detail(request, pk):
    message = get_object_or_404(MessagesAscenseursDetails, pk=pk)
    
    if request.method == 'POST':
        form = MessageDetailForm(request.POST)
        if form.is_valid():
            fields = [field for field, value in form.cleaned_data.items() if value]
    else:
        form = MessageDetailForm()
        fields = [field.verbose_name for field in MessagesAscenseursDetails._meta.get_fields() if field.name != 'N_des_messages']

    return render(request, 'accesclient/message_detail.html', {'message': message, 'form': form, 'fields': fields})



def export_messages_to_excel(request):
    user = request.user
    messages = MessagesAscenseurs.objects.filter(Destinataire=user.username)

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    # Define headers for the Excel file
    headers = [
        'Destinataire',
        'Date',
        'Message',
        'Nom',
        'Téléphone',
        'Digicode',
        'Action',
        'Société de l\'appelant',
        'Nom de l\'appelant',
        'Adresse de l\'appelant',
        'Code postal de l\'appelant',
        'Ville de l\'appelant',
        'Téléphone de l\'appelant',
        'Digicode de l\'appelant',
        'Nature de l\'appel'
    ]

    # Write headers to the first row in the worksheet
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, message in enumerate(messages, 2):  # Start from row 2 for data
        ws.cell(row=row_num, column=1, value=message.Destinataire)
        # Convert timezone-aware datetime to local timezone
        local_date = timezone.localtime(message.Date)
        # Set tzinfo=None to avoid Excel error
        ws.cell(row=row_num, column=2, value=local_date.replace(tzinfo=None))
        ws.cell(row=row_num, column=3, value=message.Message)
        ws.cell(row=row_num, column=4, value=message.Nom)
        ws.cell(row=row_num, column=5, value=message.Téléphone)
        ws.cell(row=row_num, column=6, value=message.Digicode)
        ws.cell(row=row_num, column=7, value=message.Action)
        ws.cell(row=row_num, column=8, value=message.Société_de_l_appelant)
        ws.cell(row=row_num, column=9, value=message.Nom_de_l_appelant)
        ws.cell(row=row_num, column=10, value=message.Adresse_de_l_appelant)
        ws.cell(row=row_num, column=11, value=message.Code_postal_de_l_appelant)
        ws.cell(row=row_num, column=12, value=message.Ville_de_l_appelant)
        ws.cell(row=row_num, column=13, value=message.Téléphone_de_l_appelant)
        ws.cell(row=row_num, column=14, value=message.Digicode_de_l_appelant)
        ws.cell(row=row_num, column=15, value=message.Nature_de_l_appel)

   
    filename = f"messages_{user.username}.xlsx"

    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


class AppareilView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user

        # Fetch only the Appareil records where Destinataire matches the logged-in user
        appareils_list = Appareil.objects.filter(Destinataire=user)

        # Example of excluding specific columns
        excluded_columns = ['s_Lineage', 'dateImport', 'Autres_1', 'Autres_2', 'Observations', 'Id_Societe']

        # Example of custom column names
        custom_column_names = {
            'Code_Client': 'Client Code',
            'DateCreation': 'Date of Creation',
            'Client': 'Client',
            #'Opérateur': 'Opérateur',
            'Entretien': 'Entretien',
            'Destinataire': 'Destinataire',
            #'Télésurveillance': 'Télésurveillance',
            'Adresse': 'Adresse',
            'Code_Postal': 'Code_Postal',
            'Ville': 'Ville',
            'Résidence': 'Residence',
            'Informations': 'Informations',
            'Type': 'Type',
            #'MaJFacture': 'MaJFacture',
            'Incarcération': 'Incarcération',
            #'Code_consigne': 'Code consigne',
            'Consigne_volatile': 'Consigne volatile',
            #'Utilisateur': 'Utilisateur',
            #'Opérateur2': 'Opérateur 2',
            #'MES': 'MES',
            #'RES': 'RES',
            'Phonie': 'Phonie',
            'Transmetteur': 'Transmetteur'
        }

        # Example of dynamically fetching selected columns from request.GET
        selected_columns = [field.name for field in Appareil._meta.fields if request.GET.get(field.name)]
        search_query = request.GET.get('search', '')

        # If there's a search query, filter the appareils_list
        if search_query:
            search_filter = Q()
            for field in Appareil._meta.fields:
                search_filter |= Q(**{f"{field.name}__icontains": search_query})
            appareils_list = appareils_list.filter(search_filter)

        # Paginate the appareils_list
        paginator = Paginator(appareils_list, 20)  # Show 20 appareils per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, 'accesclient/basededonnees.html', {
            'page_obj': page_obj,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names
        })

class SignUpView(generic.CreateView):
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('login')
    template_name = 'registration/signup.html'

class ArchiveMessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = ArchiveMessagesAscenseurs.objects.first()

        # Retrieve date parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        # Build cache key based on date range
        cache_key = f'archive_messages_{user.username}_{start_date_str}_{end_date_str}'
        messages_list = cache.get(cache_key)

        if messages_list is None:
            # Fetch messages for the user
            messages_list = ArchiveMessagesAscenseurs.objects.filter(Destinataire=user.username)
            
            # Determine the date range
            if start_date_str and end_date_str:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
            elif start_date_str:
                start_date = parse_date(start_date_str)
                end_date = timezone.now()
            elif end_date_str:
                start_date = timezone.make_aware(datetime.datetime.min)
                end_date = parse_date(end_date_str)
            else:
                # Default to the last 24 hours
                start_date = timezone.now() - timezone.timedelta(days=1)
                end_date = timezone.now()

            # Print debugging information
            print(f"Filtering messages from {start_date} to {end_date}")

            # Apply the date range filter
            messages_list = messages_list.filter(Date__range=[start_date, end_date])

            # Cache the results
            cache.set(cache_key, messages_list, timeout=60*15)  # Cache timeout of 15 minutes

        # Pagination
        paginator = Paginator(messages_list, 150)  # Show 10 messages per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Print debugging information
        print(f"Total messages after filtering: {paginator.count}")
        print(f"Messages on current page: {page_obj.paginator.per_page}")

        excluded_columns = [
            'Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 
            'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 
            'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages'
        ]

        # Generate custom column names dynamically
        custom_column_names = {field.name: field.verbose_name for field in ArchiveMessagesAscenseurs._meta.get_fields() if field.name not in excluded_columns}
        selected_columns = [field.name for field in ArchiveMessagesAscenseurs._meta.get_fields() if request.GET.get(field.name)]

        # Handle export functionality
        if 'export' in request.GET:
            return self.export_to_csv(page_obj, selected_columns)

        return render(request, 'accesclient/archive_messages.html', {
            'messages': messages,
            'page_obj': page_obj,
            'messages_list': page_obj,  # Pass the paginated page object
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'start_date': start_date_str,
            'end_date': end_date_str,
        })

    def export_to_csv(self, messages_list, selected_columns):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="archive_messages.csv"'
        
        writer = csv.writer(response)
        writer.writerow([col for col in selected_columns])
        
        for message in messages_list:
            row = [getattr(message, col) for col in selected_columns]
            writer.writerow(row)
        
        return response
    
def create_message(request, N_ID):
    appareil = get_object_or_404(Appareil, N_ID=N_ID)
    code_client = appareil.Code_Client
    destinataire = appareil.Destinataire

    if request.method == 'POST':
        form = MessageForm(request.POST, user=request.user)
        form = MessageForm(request.POST)
        if form.is_valid():
            
            cleaned_data = form.cleaned_data
            MessagesAscenseurs.objects.create(
                Nature_de_l_appel=cleaned_data['Nature_de_l_appel'],
                Nom_de_l_appelant=cleaned_data['Nom_de_l_appelant'],
                Société_de_l_appelant=cleaned_data['Société_de_l_appelant'],
                Téléphone_de_l_appelant=cleaned_data['Téléphone_de_l_appelant'],
                Message=cleaned_data['Message'],
                
                Destinataire=destinataire,
                N_ID=cleaned_data['N_ID']  
            )
            return redirect('http://127.0.0.1:8000/appareils/')  
    else:
        
        form = MessageForm(user=request.user)

    return render(request, 'accesclient/create_message.html', {
        'form': form,
        'code_client': code_client,
        'destinataire': destinataire,
    })

def modify_appareil(request, id):
    appareil = get_object_or_404(Appareil, pk=id)
    
    if request.method == 'POST':
        form = AppareilModificationForm(request.POST, instance=appareil)
        if form.is_valid():
            form.save()
            return redirect('http://127.0.0.1:8000/appareils/')  
    else:
        form = AppareilModificationForm(instance=appareil)
    
    return render(request, 'accesclient/modify_appareil.html', {'form': form})

def set_appareil_perdu(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Set the specified fields to "PERDU"
    appareil.Client = "PERDU"
    appareil.Opérateur = "PERDU"
    appareil.Entretien = "PERDU"
    appareil.Destinataire = "PERDU"
    appareil.Informations = "PERDU"

    # Save the changes to the database
    appareil.save()

    # Redirect to the appareils list (or wherever you want)
    return redirect('http://127.0.0.1:8000/appareils/')



def sanitize_text(text):
    if text:
        return ''.join(c for c in text if c.isprintable())
    return text
def generate_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Appareils"
    user = request.user

    headers = [
        's_Generation',
        's_GUID',
        's_Lineage',
        'N° ID',
        'DateCreation',
        'Client',
        'Opérateur',
        'Entretien',
        'Destinataire',
        'Télésurveillance',
        'Adresse',
        'Code Postal',
        'Ville',
        'Résidence',
        'Informations',
        'Code Client',
        'Type',
        'MaJFacture',
        'Incarcération',
        'Code consigne',
        'Consigne volatile',
        'Utilisateur',
        'Opérateur2',
        'dateImport',
        'MES',
        'RES',
        'Phonie',
        'Transmetteur',
        'Autres 1',
        'Autres 2',
        'Observations',
        'Id Societe'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    appareils = Appareil.objects.filter(Client=user.username)
    for row_num, appareil in enumerate(appareils.iterator(), 2):
        ws.cell(row=row_num, column=1, value=sanitize_text(appareil.s_Generation))
        ws.cell(row=row_num, column=2, value=sanitize_text(appareil.s_GUID))
        ws.cell(row=row_num, column=3, value=sanitize_text(appareil.s_Lineage))
        ws.cell(row=row_num, column=4, value=appareil.N_ID)
        if appareil.DateCreation:
            local_date_creation = timezone.localtime(appareil.DateCreation).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=5, value=local_date_creation)
        ws.cell(row=row_num, column=6, value=sanitize_text(appareil.Client))
        ws.cell(row=row_num, column=7, value=sanitize_text(appareil.Opérateur))
        ws.cell(row=row_num, column=8, value=sanitize_text(appareil.Entretien))
        ws.cell(row=row_num, column=9, value=sanitize_text(appareil.Destinataire))
        ws.cell(row=row_num, column=10, value=sanitize_text(appareil.Télésurveillance))
        ws.cell(row=row_num, column=11, value=sanitize_text(appareil.Adresse))
        ws.cell(row=row_num, column=12, value=sanitize_text(appareil.Code_Postal))
        ws.cell(row=row_num, column=13, value=sanitize_text(appareil.Ville))
        ws.cell(row=row_num, column=14, value=sanitize_text(appareil.Résidence))
        ws.cell(row=row_num, column=15, value=sanitize_text(appareil.Informations))
        ws.cell(row=row_num, column=16, value=sanitize_text(appareil.Code_Client))
        ws.cell(row=row_num, column=17, value=sanitize_text(appareil.Type))
        ws.cell(row=row_num, column=18, value=appareil.MaJFacture)
        ws.cell(row=row_num, column=19, value=sanitize_text(appareil.Incarcération))
        ws.cell(row=row_num, column=20, value=sanitize_text(appareil.Code_consigne))
        ws.cell(row=row_num, column=21, value=sanitize_text(appareil.Consigne_volatile))
        ws.cell(row=row_num, column=22, value=sanitize_text(appareil.Utilisateur))
        ws.cell(row=row_num, column=23, value=sanitize_text(appareil.Opérateur2))
        if appareil.dateImport:
            local_date_import = timezone.localtime(appareil.dateImport).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=24, value=local_date_import)
        if appareil.MES:
            local_mes = timezone.localtime(appareil.MES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=25, value=local_mes)
        if appareil.RES:
            local_res = timezone.localtime(appareil.RES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=26, value=local_res)
        ws.cell(row=row_num, column=27, value=sanitize_text(appareil.Phonie))
        ws.cell(row=row_num, column=28, value=sanitize_text(appareil.Transmetteur))
        ws.cell(row=row_num, column=29, value=sanitize_text(appareil.Autres_1))
        ws.cell(row=row_num, column=30, value=sanitize_text(appareil.Autres_2))
        ws.cell(row=row_num, column=31, value=sanitize_text(appareil.Observations))
        ws.cell(row=row_num, column=32, value=appareil.Id_Societe)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_appareils_to_excel(request):
    user = request.user
    filename = f"appareils_{user.username}.xlsx"

    response = StreamingHttpResponse(
        generate_excel(request),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def modify_autres_if_meditrax(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Redirect if the Client is not Meditrax
    if appareil.Client != "MEDITRAX":
        return redirect('http://127.0.0.1:8000/appareils/')

    if request.method == 'POST':
        form = AppareilModificationForm(request.POST, instance=appareil)
        if form.is_valid():
            # Save only the 'Autres' field
            appareil.Autres = form.cleaned_data['Autres_1']
            appareil.save()
            return redirect('http://127.0.0.1:8000/appareils/')
    else:
        form = AppareilModificationForm(instance=appareil)

    return render(request, 'accesclient/Meditrax.html', {'form': form, 'appareil': appareil})



def create_astreinte(request):
    if request.method == 'POST':
        form = AstreinteForm(request.POST)
        if form.is_valid():
            astreinte = form.save(commit=False)  
            astreinte.operator_create = request.user.username  
            astreinte.save()  
            return redirect('http://127.0.0.1:8000/create-astreinte/')  
    else:
        form = AstreinteForm()

    return render(request, 'Astreinte/astreinte_form.html', {'form': form})

def upload_excel(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            created_by = request.user.username  
            result = process_excel_file(excel_file, created_by)

            if result is True:
                messages.success(request, "Fichier traité et données insérées avec succès !")
            else:
                # If result is a list of errors, display them
                for error in result:
                    messages.error(request, error)
            return redirect("upload_excel")
    else:
        form = ExcelUploadForm()

    return render(request, "Astreinte/upload_excel.html", {"form": form})

def view_astreintes(request):
    user = request.user
    sort_order = request.GET.get('sort', 'date_debut')
    search_query = request.GET.get('q', '')
    filter_date = request.GET.get('filter_date', '')
    selected_entretien = request.GET.get('entretien', '')

    # Find all unique entretiens that start with the user's username
    # Sort to ensure base entretien (e.g., 'FCT') comes first
    available_entretiens = list(Astreinte.objects.filter(
        entretien__startswith=user.username
    ).values_list('entretien', flat=True).distinct().order_by('entretien'))

    # If no specific entretien is selected, default to the base entretien (without location)
    base_entretien = next((e for e in available_entretiens if e == user.username), None)
    if not selected_entretien:
        selected_entretien = base_entretien or available_entretiens[0] 

    # Parse dates
    try:
        search_date = datetime.strptime(search_query, '%Y-%m-%d').date()
    except ValueError:
        search_date = None

    try:
        filter_date_parsed = datetime.strptime(filter_date, '%Y-%m-%d').date()
    except ValueError:
        filter_date_parsed = None

    # Build the base query for the specific user
    filters = Q(entretien__startswith=user.username)

    # Add specific entretien filter if selected
    if selected_entretien:
        filters &= Q(entretien=selected_entretien)

    # Search query filters
    if search_query:
        search_filters = (
            Q(detail_astreinte__icontains=search_query) |
            Q(operator_create__icontains=search_query) |
            Q(media1__icontains=search_query) |
            Q(media2__icontains=search_query) |
            Q(media3__icontains=search_query) |
            Q(media4__icontains=search_query)
        )
        
        # Attempt to parse as a date if the search query looks like a date
        try:
            search_date = datetime.strptime(search_query, '%Y-%m-%d').date()
            search_filters |= Q(date_debut=search_date) | Q(date_fin=search_date)
        except ValueError:
            pass

        filters &= search_filters

    # Date filter
    if filter_date_parsed:
        filters &= Q(date_debut=filter_date_parsed)

    # Query astreintes with applied filters
    astreintes = Astreinte.objects.filter(filters).order_by(sort_order)

    return render(request, 'Astreinte/view_astreintes.html', {
        'astreintes': astreintes,
        'sort_order': sort_order,
        'search_query': search_query,
        'filter_date': filter_date,
        'selected_entretien': selected_entretien,
        'available_entretiens': available_entretiens,
    })
def delete_astreinte(request, id_astreinte):
    astreinte = get_object_or_404(Astreinte, id_astreinte=id_astreinte)
    astreinte.delete()
    messages.success(request, "L'astreinte a été supprimée avec succès.")
    return redirect(reverse('view_astreintes'))


def modify_astreinte(request, id_astreinte):
    astreinte = get_object_or_404(Astreinte, id_astreinte=id_astreinte)

    if request.method == 'POST':
        form = AstreinteForm(request.POST, instance=astreinte)
        if form.is_valid():
            form.save()
            messages.success(request, "L'astreinte a été modifiée avec succès.")
            return redirect(reverse('view_astreintes'))
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = AstreinteForm(instance=astreinte)

    return render(request, 'Astreinte/modify_astreinte.html', {'form': form, 'astreinte': astreinte})