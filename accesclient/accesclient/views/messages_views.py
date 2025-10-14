# messages_views.py
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.core.cache import cache
from openpyxl import Workbook
import csv
from datetime import datetime

from ..models import MessagesAscenseurs, MessagesAscenseursDetails, ArchiveMessagesAscenseurs, Appareil
from ..forms import MessageDetailForm, MessageForm


class MessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = MessagesAscenseursDetails.objects.first()
        
        # Filter messages based on user type
        if Appareil.objects.filter(Client=user.first_name).exists():
            messages_list = MessagesAscenseursDetails.objects.filter(Destinataire=user.first_name)
        else:
            messages_list = MessagesAscenseursDetails.objects.filter(entretien=user.first_name)
        
        # Get unique "Entretien" values for the user
        entretiens = messages_list.values_list('entretien', flat=True).distinct()
        
        # Get the selected "Entretien" from GET parameters
        selected_entretien = request.GET.get('entretien')

        # Filter messages based on selected "Entretien"
        if selected_entretien:
            messages_list = messages_list.filter(entretien=selected_entretien)
        
        excluded_columns = ['Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages']  
        custom_column_names = {
            'Date': 'Date du message',
            'Nature_de_l_appel': 'Type d\'appel',
            'code_client': 'N°APP',
            'Résidence': 'Coordonnées du Site',
            'Message': 'Message',
            'Action': 'Action',
            'Nom': 'Nom',
            'Société_de_l_appelant': 'Coord. de l\'appelant',
            'Nom_de_l_appelant': 'Nom de l\'appelant',
            'Téléphone_de_l_appelant': 'Téléphone ',
            'Adresse_de_l_appelant': 'Adresse de l\'appelant',
            'Code_postal_de_l_appelant': 'Code postal de l\'appelant',
            'Ville_de_l_appelant': 'Ville de l\'appelant',
        }

        # Combine the content of the three columns into one called 'Résidence'
        for message in messages_list:
            message.Résidence = f"{message.Adresse}, {message.Code_Postal}, {message.ville}"
        
        selected_columns = [field_name for field_name in messages.get_fields() if request.GET.get(field_name)]

        return render(request, 'accesclient/mesasc.html', {
            'messages': messages,
            'messages_list': messages_list,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'entretiens': entretiens,
            'selected_entretien': selected_entretien,
        })


class ArchiveMessagesView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        messages = ArchiveMessagesAscenseurs.objects.first()

        # Retrieve date parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        # Build cache key based on date range
        cache_key = f'archive_messages_{user.username}_{start_date_str}_{end_date_str}'
        messages_list = cache.get(cache_key)

        if messages_list is None:
            # Fetch messages for the user based on user type
            if Appareil.objects.filter(Client=user.first_name).exists():
                messages_list = ArchiveMessagesAscenseurs.objects.filter(Destinataire=user.first_name)
            else:
                messages_list = ArchiveMessagesAscenseurs.objects.filter(entretien=user.first_name)
            
            # Determine the date range
            if start_date_str and end_date_str:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
            elif start_date_str:
                start_date = parse_date(start_date_str)
                end_date = timezone.now()
            elif end_date_str:
                start_date = timezone.make_aware(datetime.datetime.min)
                end_date = parse_date(end_date_str)
            else:
                # Default to the last 24 hours
                start_date = timezone.now() - timezone.timedelta(days=1)
                end_date = timezone.now()

            # Print debugging information
            print(f"Filtering messages from {start_date} to {end_date}")

            # Apply the date range filter
            messages_list = messages_list.filter(Date__range=[start_date, end_date])

            # Cache the results
            cache.set(cache_key, messages_list, timeout=60*15)  # Cache timeout of 15 minutes

        # Pagination
        paginator = Paginator(messages_list, 150)  # Show 150 messages per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Print debugging information
        print(f"Total messages after filtering: {paginator.count}")
        print(f"Messages on current page: {page_obj.paginator.per_page}")

        excluded_columns = [
            'Stocké', 'Incarcération', 'Opérateur', 'Confirmation', 
            'ConfIncar', 'ConfIncar2', 'Commentaires', 'Autres1', 
            'Autres2', 'Etat', 'Téléphone_2', 'N_ID', 'N_des_messages',
            'Destinataire'
        ]

        # Generate custom column names dynamically
        custom_column_names = {field.name: field.verbose_name for field in ArchiveMessagesAscenseurs._meta.get_fields() if field.name not in excluded_columns}
        selected_columns = [field.name for field in ArchiveMessagesAscenseurs._meta.get_fields() if request.GET.get(field.name)]

        # Handle export functionality
        if 'export' in request.GET:
            return self.export_to_csv(page_obj, selected_columns)

        return render(request, 'accesclient/archive_messages.html', {
            'messages': messages,
            'page_obj': page_obj,
            'messages_list': page_obj,  # Pass the paginated page object
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names,
            'start_date': start_date_str,
            'end_date': end_date_str,
        })

    def export_to_csv(self, messages_list, selected_columns):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="archive_messages.csv"'
        
        writer = csv.writer(response)
        writer.writerow([col for col in selected_columns])
        
        for message in messages_list:
            row = [getattr(message, col) for col in selected_columns]
            writer.writerow(row)
        
        return response


def messages_list(request):
    user = request.user
    messages = MessagesAscenseursDetails.objects.filter(Destinataire=user.username)
    
    date_str = request.GET.get('date')
    if date_str:
        try:
            date_obj = parse_date(date_str)
            if date_obj:
                messages = messages.filter(Date__date=date_obj)
        except ValueError:
            pass
    
    return render(request, 'accesclient/messages_list.html', {'messages': messages})


def message_detail(request, pk):
    message = get_object_or_404(MessagesAscenseursDetails, pk=pk)
    
    if request.method == 'POST':
        form = MessageDetailForm(request.POST)
        if form.is_valid():
            fields = [field for field, value in form.cleaned_data.items() if value]
    else:
        form = MessageDetailForm()
        fields = [field.verbose_name for field in MessagesAscenseursDetails._meta.get_fields() if field.name != 'N_des_messages']

    return render(request, 'accesclient/message_detail.html', {'message': message, 'form': form, 'fields': fields})


def create_message(request, N_ID):
    appareil = get_object_or_404(Appareil, N_ID=N_ID)
    code_client = appareil.Code_Client
    destinataire = appareil.Destinataire

    if request.method == 'POST':
        form = MessageForm(request.POST, user=request.user)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            MessagesAscenseurs.objects.create(
                Nature_de_l_appel=cleaned_data['Nature_de_l_appel'],
                Nom_de_l_appelant=cleaned_data['Nom_de_l_appelant'],
                Société_de_l_appelant=cleaned_data['Société_de_l_appelant'],
                Téléphone_de_l_appelant=cleaned_data['Téléphone_de_l_appelant'],
                Message=cleaned_data['Message'],
                Destinataire=destinataire,
                N_ID=N_ID  # Use the N_ID from URL parameter directly
            )
            return redirect('http://127.0.0.1:8000/appareils/')  
    else:
        # Initialize form with N_ID value
        form = MessageForm(user=request.user, initial={'N_ID': N_ID})

    return render(request, 'accesclient/create_message.html', {
        'form': form,
        'code_client': code_client,
        'destinataire': destinataire,
    })


def export_messages_to_excel(request):
    user = request.user
    messages = MessagesAscenseurs.objects.filter(Destinataire=user.username)

    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    # Define headers for the Excel file
    headers = [
        'Destinataire',
        'Date',
        'Message',
        'Nom',
        'Téléphone',
        'Digicode',
        'Action',
        'Société de l\'appelant',
        'Nom de l\'appelant',
        'Adresse de l\'appelant',
        'Code postal de l\'appelant',
        'Ville de l\'appelant',
        'Téléphone de l\'appelant',
        'Digicode de l\'appelant',
        'Nature de l\'appel'
    ]

    # Write headers to the first row in the worksheet
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, message in enumerate(messages, 2):  # Start from row 2 for data
        ws.cell(row=row_num, column=1, value=message.Destinataire)
        # Convert timezone-aware datetime to local timezone
        local_date = timezone.localtime(message.Date)
        # Set tzinfo=None to avoid Excel error
        ws.cell(row=row_num, column=2, value=local_date.replace(tzinfo=None))
        ws.cell(row=row_num, column=3, value=message.Message)
        ws.cell(row=row_num, column=4, value=message.Nom)
        ws.cell(row=row_num, column=5, value=message.Téléphone)
        ws.cell(row=row_num, column=6, value=message.Digicode)
        ws.cell(row=row_num, column=7, value=message.Action)
        ws.cell(row=row_num, column=8, value=message.Société_de_l_appelant)
        ws.cell(row=row_num, column=9, value=message.Nom_de_l_appelant)
        ws.cell(row=row_num, column=10, value=message.Adresse_de_l_appelant)
        ws.cell(row=row_num, column=11, value=message.Code_postal_de_l_appelant)
        ws.cell(row=row_num, column=12, value=message.Ville_de_l_appelant)
        ws.cell(row=row_num, column=13, value=message.Téléphone_de_l_appelant)
        ws.cell(row=row_num, column=14, value=message.Digicode_de_l_appelant)
        ws.cell(row=row_num, column=15, value=message.Nature_de_l_appel)

    filename = f"messages_{user.username}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
