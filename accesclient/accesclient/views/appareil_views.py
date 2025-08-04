# appareil_views.py
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q
from openpyxl import Workbook
from io import BytesIO

from ..models import Appareil
from ..forms import AppareilModificationForm


class AppareilView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user

        # Fetch Appareil records based on user type
        if user.is_superuser:
            appareils_list = Appareil.objects.filter(Destinataire=user)
        else:
            appareils_list = Appareil.objects.filter(Entretien=user.first_name)

        # Example of excluding specific columns
        excluded_columns = ['s_Lineage', 'dateImport', 'Autres_1', 'Autres_2', 'Observations', 'Id_Societe']

        # Example of custom column names
        custom_column_names = {
            'Code_Client': 'Client Code',
            'DateCreation': 'Date of Creation',
            'Client': 'Client',
            'Entretien': 'Entretien',
            'Destinataire': 'Destinataire',
            'Adresse': 'Adresse',
            'Code_Postal': 'Code_Postal',
            'Ville': 'Ville',
            'Résidence': 'Residence',
            'Informations': 'Informations',
            'Type': 'Type',
            'Incarcération': 'Incarcération',
            'Consigne_volatile': 'Consigne volatile',
            'Phonie': 'Phonie',
            'Transmetteur': 'Transmetteur'
        }

        # Example of dynamically fetching selected columns from request.GET
        selected_columns = [field.name for field in Appareil._meta.fields if request.GET.get(field.name)]
        search_query = request.GET.get('search', '')

        # If there's a search query, filter the appareils_list
        if search_query:
            search_filter = Q()
            for field in Appareil._meta.fields:
                search_filter |= Q(**{f"{field.name}__icontains": search_query})
            appareils_list = appareils_list.filter(search_filter)

        # Paginate the appareils_list
        paginator = Paginator(appareils_list, 20)  # Show 20 appareils per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, 'accesclient/basededonnees.html', {
            'page_obj': page_obj,
            'selected_columns': selected_columns,
            'excluded_columns': excluded_columns,
            'custom_column_names': custom_column_names
        })


def modify_appareil(request, id):
    appareil = get_object_or_404(Appareil, pk=id)
    
    if request.method == 'POST':
        form = AppareilModificationForm(request.POST, instance=appareil)
        if form.is_valid():
            form.save()
            return redirect('http://127.0.0.1:8000/appareils/')  
    else:
        form = AppareilModificationForm(instance=appareil)
    
    return render(request, 'accesclient/modify_appareil.html', {'form': form})


def set_appareil_perdu(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Set the specified fields to "PERDU"
    appareil.Client = "PERDU"
    appareil.Opérateur = "PERDU"
    appareil.Entretien = "PERDU"
    appareil.Destinataire = "PERDU"
    appareil.Informations = "PERDU"

    # Save the changes to the database
    appareil.save()

    # Redirect to the appareils list (or wherever you want)
    return redirect('http://127.0.0.1:8000/appareils/')


def modify_autres_if_meditrax(request, id):
    appareil = get_object_or_404(Appareil, pk=id)

    # Redirect if the Client is not Meditrax
    if appareil.Client != "MEDITRAX":
        return redirect('http://127.0.0.1:8000/appareils/')

    if request.method == 'POST':
        form = AppareilModificationForm(request.POST, instance=appareil)
        if form.is_valid():
            # Save only the 'Autres' field
            appareil.Autres = form.cleaned_data['Autres_1']
            appareil.save()
            return redirect('http://127.0.0.1:8000/appareils/')
    else:
        form = AppareilModificationForm(instance=appareil)

    return render(request, 'accesclient/Meditrax.html', {'form': form, 'appareil': appareil})


def sanitize_text(text):
    if text:
        return ''.join(c for c in text if c.isprintable())
    return text


def generate_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Appareils"
    user = request.user

    headers = [
        's_Generation',
        's_GUID',
        's_Lineage',
        'N° ID',
        'DateCreation',
        'Client',
        'Opérateur',
        'Entretien',
        'Destinataire',
        'Télésurveillance',
        'Adresse',
        'Code Postal',
        'Ville',
        'Résidence',
        'Informations',
        'Code Client',
        'Type',
        'MaJFacture',
        'Incarcération',
        'Code consigne',
        'Consigne volatile',
        'Utilisateur',
        'Opérateur2',
        'dateImport',
        'MES',
        'RES',
        'Phonie',
        'Transmetteur',
        'Autres 1',
        'Autres 2',
        'Observations',
        'Id Societe'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    appareils = Appareil.objects.filter(Client=user.username)
    for row_num, appareil in enumerate(appareils.iterator(), 2):
        ws.cell(row=row_num, column=1, value=sanitize_text(appareil.s_Generation))
        ws.cell(row=row_num, column=2, value=sanitize_text(appareil.s_GUID))
        ws.cell(row=row_num, column=3, value=sanitize_text(appareil.s_Lineage))
        ws.cell(row=row_num, column=4, value=appareil.N_ID)
        if appareil.DateCreation:
            local_date_creation = timezone.localtime(appareil.DateCreation).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=5, value=local_date_creation)
        ws.cell(row=row_num, column=6, value=sanitize_text(appareil.Client))
        ws.cell(row=row_num, column=7, value=sanitize_text(appareil.Opérateur))
        ws.cell(row=row_num, column=8, value=sanitize_text(appareil.Entretien))
        ws.cell(row=row_num, column=9, value=sanitize_text(appareil.Destinataire))
        ws.cell(row=row_num, column=10, value=sanitize_text(appareil.Télésurveillance))
        ws.cell(row=row_num, column=11, value=sanitize_text(appareil.Adresse))
        ws.cell(row=row_num, column=12, value=sanitize_text(appareil.Code_Postal))
        ws.cell(row=row_num, column=13, value=sanitize_text(appareil.Ville))
        ws.cell(row=row_num, column=14, value=sanitize_text(appareil.Résidence))
        ws.cell(row=row_num, column=15, value=sanitize_text(appareil.Informations))
        ws.cell(row=row_num, column=16, value=sanitize_text(appareil.Code_Client))
        ws.cell(row=row_num, column=17, value=sanitize_text(appareil.Type))
        ws.cell(row=row_num, column=18, value=appareil.MaJFacture)
        ws.cell(row=row_num, column=19, value=sanitize_text(appareil.Incarcération))
        ws.cell(row=row_num, column=20, value=sanitize_text(appareil.Code_consigne))
        ws.cell(row=row_num, column=21, value=sanitize_text(appareil.Consigne_volatile))
        ws.cell(row=row_num, column=22, value=sanitize_text(appareil.Utilisateur))
        ws.cell(row=row_num, column=23, value=sanitize_text(appareil.Opérateur2))
        if appareil.dateImport:
            local_date_import = timezone.localtime(appareil.dateImport).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=24, value=local_date_import)
        if appareil.MES:
            local_mes = timezone.localtime(appareil.MES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=25, value=local_mes)
        if appareil.RES:
            local_res = timezone.localtime(appareil.RES).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=26, value=local_res)
        ws.cell(row=row_num, column=27, value=sanitize_text(appareil.Phonie))
        ws.cell(row=row_num, column=28, value=sanitize_text(appareil.Transmetteur))
        ws.cell(row=row_num, column=29, value=sanitize_text(appareil.Autres_1))
        ws.cell(row=row_num, column=30, value=sanitize_text(appareil.Autres_2))
        ws.cell(row=row_num, column=31, value=sanitize_text(appareil.Observations))
        ws.cell(row=row_num, column=32, value=appareil.Id_Societe)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_appareils_to_excel(request):
    user = request.user
    filename = f"appareils_{user.username}.xlsx"

    response = StreamingHttpResponse(
        generate_excel(request),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
